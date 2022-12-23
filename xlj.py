from openpyxl import Workbook  # 「pip install openpyxl」でインストールしておく

######################################## 
import os
import random
import string
import time

######################################## 並列処理
from concurrent.futures import \
    ProcessPoolExecutor  # /マルチスレッド　プロセスが併用　3.2から始まった
from concurrent.futures import ThreadPoolExecutor  # /マルチスレッド　プロセスが併用　3.2から始まった
from multiprocessing import Pipe, Process
######################################## TIKTOK API
from re import A
from this import d
from TikTokLive import TikTokLiveClient
from TikTokLive.types.events import CommentEvent, ConnectEvent,ViewerCountUpdateEvent,ViewerCountUpdateEvent,EnvelopeEvent,JoinEvent
import os

########################################
#######################################　ローカル・グルーバル関数定義 global likeswitch ライク判定のスイッチ
likeswitch = 0
likecount = 0
baniracounts = 0
connectcount = 0 #接続前の文字のために待機　スイッチ
connectcount2 = 0 #読まれるランダム　スイッチ
tiktoksheet1 = 4 #なん行目から書き込みするか
coment_count = 0 #コメントの書き込まれた回数
view_now =0
index = 2 
rown=1
######################################## #現在時刻をコメント
import datetime
t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
dt_now = datetime.datetime.now()
dt_no =  '{:%Y%m%d%H%M%S}'.format(now)
######################################## 

mastername = "@punks5d"


client = TikTokLiveClient(f"{mastername}", **{
})


wbname = (f"{mastername}"+dt_no+'_join.xlsx')
#@punks5d
# ワークブックの新規作成と保存
wb = Workbook()

wb.save(wbname)

# ワークブックの読み込み
from openpyxl import load_workbook
wb = load_workbook(wbname)
ws = wb['Sheet']  # ワークシートを指定
sheet = wb.worksheets[0]


######################################## #クラインと接続
@client.on("connect")
async def on_connect(_: ConnectEvent):
    global connectcount
    print("Connected to Room ID:", client.room_id)
    print(mastername)    
    connectcount += 1

######################################## #入ったら

@client.on("join")
async def on_join(event: JoinEvent):
    #print("Someone joined the stream!")
    global coment_count
    print("join" +f"{coment_count}")
    # セルへ書き込む
    dt_now = datetime.datetime.now()
    sheet.append([f"{event.user.nickname}",f"{event.user.uniqueId}","","join"+f"{coment_count}",f"{dt_now}",f"{mastername}","join"])
    wb.save(wbname)
    coment_count +=1

if __name__ == '__main__':    
    with ProcessPoolExecutor(max_workers=6) as executor:
        executor.submit(on_join)
    client.run()



