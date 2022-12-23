from openpyxl import Workbook 

######################################## 
import os
import random
import string
import time

######################################## 
from concurrent.futures import \
    ProcessPoolExecutor  
from concurrent.futures import ThreadPoolExecutor 
from multiprocessing import Pipe, Process
######################################## 
from re import A
from this import d
from TikTokLive import TikTokLiveClient
from TikTokLive.types.events import CommentEvent, ConnectEvent,ViewerCountUpdateEvent,ViewerCountUpdateEvent,EnvelopeEvent,JoinEvent,GiftEvent
import os

########################################
#######################################　
likeswitch = 0
likecount = 0
baniracounts = 0
connectcount = 0 #接続前の文字のために待機　スイッチ
connectcount2 = 0 #読まれるランダム　スイッチ
tiktoksheet1 = 4 #なん行目から書き込みするか
coment_count = 0 #コメントの書き込まれた回数
view_now =0
index = 2 
rown=1


######################################## 
import datetime
t_delta = datetime.timedelta(hours=9)
JST = datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)

dt_now = datetime.datetime.now()
dt_no =  '{:%Y%m%d%H%M%S}'.format(now)

######################################## 

######################################## ここのIDをデータを取りたい人に書き換える　このプログラムがある場所に保存される

mastername = "@punks5d"


client = TikTokLiveClient(f"{mastername}", **{
})


#@punks5d
wbname = (f"{mastername}"+dt_no+'_comment.xlsx')
# ワークブックの新規作成と保存
wb = Workbook()
wb.save(wbname)
# ワークブックの読み込み
from openpyxl import load_workbook
wb = load_workbook(wbname)
ws = wb['Sheet']  # ワークシートを指定
sheet = wb.worksheets[0]

######################################## 
@client.on("connect")
async def on_connect(_: ConnectEvent):
    global connectcount
    print("Connected to Room ID:", client.room_id)
    print(mastername)
    connectcount += 1

######################################## 宝箱を投げられた数と投げた人をエクセルに記載して出力

@client.on("envelope")
async def on_connect(event: EnvelopeEvent):

    dt_now = datetime.datetime.now()
    print("coin")
    sheet.append([f"{event.treasureBoxUser.nickname}",f"{event.treasureBoxUser.uniqueId}",f"{event.treasureBoxData.coins}","coin",f"{dt_now}",f"{mastername}","coin"])

    wb.save(wbname)
##################################### コメントをのこした人をエクセルに記載して出力

@client.on("comment")
async def on_comment(event: CommentEvent):
    global connectcount
    global connectcount2
    global tiktoksheet1
    global index
    global coment_count
    global view_now
    global rown
    if connectcount > 0:
        if connectcount2 == 0:
            print(f"{event.comment}")
            dt_now = datetime.datetime.now()
            sheet.append([f"{event.user.nickname}",f"{event.user.uniqueId}",f"{event.comment}",f"{coment_count}",f"{dt_now}",f"{mastername}","comment"])
            wb.save(wbname)
            
    connectcount2 =random.randint(0, 0) 


##################################### ギフトをなげた人をエクセルに記載して出力

@client.on("gift")
async def on_gift(event: GiftEvent):
            dt_now = datetime.datetime.now()
            sheet.append([f"{event.user.nickname}",f"{event.user.uniqueId}",f"{event.gift.extended_gift.name}",f"{event.gift.repeat_count}",f"{dt_now}",f"{mastername}","gift"])
            wb.save(wbname)

##################################### 

if __name__ == '__main__':

    
    with ProcessPoolExecutor(max_workers=6) as executor:
        #executor.submit(on_comment)
        executor.submit(on_comment)
        executor.submit(on_connect)

    client.run()



